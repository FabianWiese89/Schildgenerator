import os
import openpyxl
import qrcode

from PIL import Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas

from src.layouts import (
    SINGLE_LABEL_LAYOUT,
    FOUR_LINE_LAYOUT,
    FIVE_LINE_LAYOUT,
    TEXT_SIGN_A4_LANDSCAPE_LAYOUT,
)
from src.paths import resource_path


def fit_text_to_safe_area(pdf, lines, safe_area_width, safe_area_height, layout):
    font_size = layout["max_font_size"]

    while font_size >= layout["min_font_size"]:
        line_spacing = font_size * 1.2
        total_text_height = len(lines) * line_spacing

        longest_line_width = max(
            pdf.stringWidth(
                line,
                layout["font_name"],
                font_size
            )
            for line in lines
        )

        if (
            longest_line_width <= safe_area_width
            and total_text_height <= safe_area_height
        ):
            return font_size

        font_size -= 2

    return layout["min_font_size"]

def split_text_into_lines(pdf, text, safe_area_width, font_name, font_size):
    words = text.split()

    lines = []
    current_line = ""

    for word in words:
        if current_line == "":
            test_line = word
        else:
            test_line = current_line + " " + word

        test_width = pdf.stringWidth(
            test_line,
            font_name,
            font_size
        )

        if test_width <= safe_area_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)

            current_line = word

    if current_line:
        lines.append(current_line)

    return lines

def generate_pdf_einzeln(text_vis, qr_data, output, zeilen):
    if zeilen == 4:
        layout_profile = FOUR_LINE_LAYOUT
        text_y = landscape(A4)[1] - 100 - 56.6929 - 28.3465
    else:
        layout_profile = FIVE_LINE_LAYOUT
        text_y = landscape(A4)[1] - 100 - 56.6929

    font_sizes = layout_profile["font_sizes"]

    border_color = colors.Color(0 / 255, 112 / 255, 60 / 255)
    pdf = canvas.Canvas(output, pagesize=landscape(A4))
    width, height = landscape(A4)

    margin = SINGLE_LABEL_LAYOUT["margin"]
    pdf.setStrokeColor(border_color)
    pdf.setLineWidth(SINGLE_LABEL_LAYOUT["border_width"])
    pdf.rect(margin, margin, width - 2 * margin, height - 2 * margin)

    text_x = 50
    left_width = width / 2 - 50
    lines = text_vis.split()

    for i, line in enumerate(lines):
        size = font_sizes[min(i, len(font_sizes) - 1)]
        pdf.setFont("Times-Bold", size)

        tw = pdf.stringWidth(line, "Times-Bold", size)

        if zeilen == 5 and (i == len(lines) - 1 or i == 4):
            text_y += 28.3465

        pdf.drawString(text_x + (left_width - tw) / 2, text_y, line)
        text_y -= size + 22.6772

    footer_text = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.0"
    pdf.setFont("Times-Roman", 12)
    pdf.drawString(50, 30, footer_text)

    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=1,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")
    tmp_qr = "temp_qr_single.png"
    qr_img.save(tmp_qr)

    qr_max = min(width / 2 - 50, height - 200)
    qr_img = Image.open(tmp_qr).resize(
        (int(qr_max), int(qr_max)),
        Image.LANCZOS,
    )

    qr_x = width / 2 + 50 - 42.5197 - 56.6929
    qr_y = (height - qr_max) / 2 - 28.3465

    pdf.drawImage(tmp_qr, qr_x, qr_y, qr_max, qr_max)
    os.remove(tmp_qr)

    logo_file = resource_path("assets/Logo_White.jpeg")
    logo_img = Image.open(logo_file)

    logo_w = qr_max * 0.5
    logo_h = logo_w * logo_img.size[1] / logo_img.size[0]

    logo_img = logo_img.resize((int(logo_w), int(logo_h)), Image.LANCZOS)

    tmp_logo = "temp_logo_single.png"
    logo_img.save(tmp_logo)

    pdf.drawImage(
        tmp_logo,
        qr_x + (qr_max - logo_w) / 2,
        qr_y + qr_max + 10,
        logo_w,
        logo_h,
    )
    os.remove(tmp_logo)

    pdf.showPage()
    pdf.save()


def generate_text_sign_pdf(title_text, text, output, show_safe_area=False):
    layout = TEXT_SIGN_A4_LANDSCAPE_LAYOUT

    pdf = canvas.Canvas(output, pagesize=landscape(A4))
    width, height = landscape(A4)

    border_color = colors.Color(0 / 255, 112 / 255, 60 / 255)

    margin = layout["margin"]
    pdf.setStrokeColor(border_color)
    pdf.setLineWidth(layout["border_width"])
    pdf.rect(margin, margin, width - 2 * margin, height - 2 * margin)

    logo_file = resource_path(layout["logo_path"])
    logo_img = Image.open(logo_file)

    logo_w = layout["logo_max_width"]
    logo_h = logo_w * logo_img.size[1] / logo_img.size[0]

    logo_x = width - logo_w - layout["logo_padding_x"] - margin
    logo_y = height - logo_h - layout["logo_padding_y"] - margin

    tmp_logo = "temp_text_logo.png"
    logo_img.save(tmp_logo)

    pdf.drawImage(tmp_logo, logo_x, logo_y, logo_w, logo_h)
    os.remove(tmp_logo)

    pdf.setFont(layout["title_font_name"], layout["title_font_size"])

    title_x = margin + layout["title_area_padding"]
    title_y = (
        height
        - margin
        - layout["title_area_padding"]
        - layout["title_font_size"]
    )

    pdf.drawString(title_x, title_y, title_text)

    safe_area_x = margin + layout["text_area_left_padding"]
    safe_area_y = margin + layout["text_area_bottom_padding"]

    safe_area_width = (
        width
        - (2 * margin)
        - layout["text_area_left_padding"]
        - layout["text_area_right_padding"]
    )

    safe_area_height = (
        height
        - (2 * margin)
        - layout["title_area_height"]
        - layout["text_area_top_padding"]
        - layout["text_area_bottom_padding"]
    )

    if show_safe_area:
        pdf.setStrokeColorRGB(1, 0, 0)
        pdf.rect(
            safe_area_x,
            safe_area_y,
            safe_area_width,
            safe_area_height,
        )

    lines = split_text_into_lines(
        pdf,
        text,
        safe_area_width,
        layout["font_name"],
        layout["max_font_size"],
    )   

    longest_line = max(lines, key=len)

    font_size = fit_text_to_safe_area(
        pdf,
        lines,
        safe_area_width,
        safe_area_height,
        layout,
    )

    pdf.setFont(layout["font_name"], font_size)

    line_spacing = font_size * 1.2
    total_text_height = len(lines) * line_spacing

    start_y = safe_area_y + (safe_area_height + total_text_height) / 2 - font_size

    for line in lines:
        line_width = pdf.stringWidth(
            line,
            layout["font_name"],
            font_size,
        )

        line_x = safe_area_x + (safe_area_width - line_width) / 2

        pdf.drawString(
            line_x,
            start_y,
            line
        )

        start_y -= line_spacing

    pdf.showPage()
    pdf.save()
    
def generate_batch_pdf_4(excel, output):
        wb = openpyxl.load_workbook(excel)
        sheet = wb.active
        font_sizes = [60, 70, 75, 105]
        border_color = colors.Color(0/255,112/255,60/255)
        pdf = canvas.Canvas(output, pagesize=landscape(A4))
        width, height = landscape(A4)
        for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
            pdf.setStrokeColor(border_color)
            pdf.setLineWidth(14.1732)
            margin = 28.3465*2
            pdf.rect(margin, margin, width-2*margin, height-2*margin)
            text_y = height-100-56.6929-28.3465
            text_x = 50
            left_width = width/2-50
            if row[0]:
                lines = str(row[0]).split()
                for i, line in enumerate(lines):
                    size = font_sizes[min(i,len(font_sizes)-1)]
                    pdf.setFont("Times-Bold", size)
                    tw = pdf.stringWidth(line, "Times-Bold", size)
                    pdf.drawString(text_x+(left_width-tw)/2, text_y, line)
                    text_y -= size+22.6772
            footer_text = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.0"
            pdf.setFont("Times-Roman",12)
            pdf.drawString(50,30, footer_text)
            qr_data = row[1] if row[1] else "https://example.com"
            qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=1)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            tmp_qr = f"temp_qr_{idx}.png"; qr_img.save(tmp_qr)
            qr_max = min(width/2-50, height-200)
            qr_img = Image.open(tmp_qr).resize((int(qr_max),int(qr_max)), Image.LANCZOS)
            qr_x = width/2+50-42.5197-56.6929; qr_y = (height-qr_max)/2-28.3465
            pdf.drawImage(tmp_qr, qr_x, qr_y, qr_max, qr_max)
            os.remove(tmp_qr)
            logo_file = resource_path("assets/Logo_White.jpeg")
            logo_img = Image.open(logo_file)
            logo_w=qr_max*0.5; logo_h=logo_w*logo_img.size[1]/logo_img.size[0]
            logo_img=logo_img.resize((int(logo_w),int(logo_h)), Image.LANCZOS)
            tmp_logo=f"temp_logo_{idx}.png"; logo_img.save(tmp_logo)
            pdf.drawImage(tmp_logo, qr_x+(qr_max-logo_w)/2, qr_y+qr_max+10, logo_w, logo_h)
            os.remove(tmp_logo)
            pdf.showPage()
        pdf.save()
        
def generate_batch_pdf_5(excel, output):
    wb = openpyxl.load_workbook(excel)
    sheet = wb.active
    font_sizes = [40, 50, 75, 105, 80]
    border_color = colors.Color(0/255,112/255,60/255)
    pdf = canvas.Canvas(output, pagesize=landscape(A4))
    width, height = landscape(A4)
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        pdf.setStrokeColor(border_color)
        pdf.setLineWidth(14.1732)
        margin = 28.3465*2
        pdf.rect(margin, margin, width-2*margin, height-2*margin)
        text_y = height-100-56.6929
        text_x = 50
        left_width = width/2-50
        if row[0]:
            lines = str(row[0]).split()
            for i, line in enumerate(lines):
                size = font_sizes[min(i,len(font_sizes)-1)]
                pdf.setFont("Times-Bold", size)
                tw = pdf.stringWidth(line, "Times-Bold", size)
                if i==len(lines)-1 or i==4:
                    text_y += 28.3465
                pdf.drawString(text_x+(left_width-tw)/2, text_y, line)
                text_y -= size+22.6772
        footer_text = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.0"
        pdf.setFont("Times-Roman",12)
        pdf.drawString(50,30, footer_text)
        qr_data = row[1] if row[1] else "https://example.com"
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=1)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        tmp_qr = f"temp_qr_{idx}.png"; qr_img.save(tmp_qr)
        qr_max = min(width/2-50, height-200)
        qr_img = Image.open(tmp_qr).resize((int(qr_max),int(qr_max)), Image.LANCZOS)
        qr_x = width/2+50-42.5197-56.6929; qr_y = (height-qr_max)/2-28.3465
        pdf.drawImage(tmp_qr, qr_x, qr_y, qr_max, qr_max)
        os.remove(tmp_qr)
        logo_file = resource_path("assets/Logo_White.jpeg")
        logo_img = Image.open(logo_file)
        logo_w=qr_max*0.5; logo_h=logo_w*logo_img.size[1]/logo_img.size[0]
        logo_img=logo_img.resize((int(logo_w),int(logo_h)), Image.LANCZOS)
        tmp_logo=f"temp_logo_{idx}.png"; logo_img.save(tmp_logo)
        pdf.drawImage(tmp_logo, qr_x+(qr_max-logo_w)/2, qr_y+qr_max+10, logo_w, logo_h)
        os.remove(tmp_logo)
        pdf.showPage()
    pdf.save()
        
