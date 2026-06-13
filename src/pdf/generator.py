import zipfile

import openpyxl
import qrcode

from PIL import Image
from openpyxl.utils.exceptions import InvalidFileException
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from src.config.layouts import (
    FIVE_LINE_LAYOUT,
    FOUR_LINE_LAYOUT,
    SINGLE_LABEL_LAYOUT,
    TEXT_SIGN_A4_LANDSCAPE_LAYOUT,
)
from src.utils import (
    ensure_excel_input_path,
    ensure_pdf_output_path,
    ensure_required_file,
    resource_path,
)

FOOTER_TEXT = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.2"
LOGO_WHITE_PATH = "assets/Logo_White.jpeg"
BORDER_COLOR = colors.Color(0 / 255, 112 / 255, 60 / 255)


def _create_pdf_canvas(output_path):
    pdf_path = ensure_pdf_output_path(output_path)
    return canvas.Canvas(str(pdf_path), pagesize=landscape(A4))


def _load_logo(relative_path=LOGO_WHITE_PATH):
    logo_path = ensure_required_file(
        resource_path(relative_path),
        "Die Logo-Datei",
    )

    try:
        return Image.open(logo_path)
    except OSError as exc:
        raise OSError(
            f"Die Logo-Datei konnte nicht geöffnet werden: {logo_path}\n"
            "Bitte prüfe, ob die Datei vorhanden und nicht beschädigt ist."
        ) from exc


def _draw_border(pdf, width, height, margin, border_width):
    pdf.setStrokeColor(BORDER_COLOR)
    pdf.setLineWidth(border_width)
    pdf.rect(margin, margin, width - 2 * margin, height - 2 * margin)


def _draw_footer(pdf):
    pdf.setFont("Times-Roman", 12)
    pdf.drawString(50, 30, FOOTER_TEXT)


def _create_qr_image(qr_data):
    if not qr_data or not str(qr_data).strip():
        raise ValueError("Für den QR-Code fehlt der Inhalt.")

    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=1,
    )
    qr.add_data(str(qr_data).strip())
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white").convert("RGB")


def _draw_qr_and_logo(pdf, width, height, qr_data):
    qr_max = min(width / 2 - 50, height - 200)
    qr_img = _create_qr_image(qr_data).resize(
        (int(qr_max), int(qr_max)),
        Image.LANCZOS,
    )

    qr_x = width / 2 + 50 - 42.5197 - 56.6929
    qr_y = (height - qr_max) / 2 - 28.3465

    pdf.drawImage(ImageReader(qr_img), qr_x, qr_y, qr_max, qr_max)

    logo_img = _load_logo()
    logo_w = qr_max * 0.5
    logo_h = logo_w * logo_img.size[1] / logo_img.size[0]
    logo_img = logo_img.resize((int(logo_w), int(logo_h)), Image.LANCZOS)

    pdf.drawImage(
        ImageReader(logo_img),
        qr_x + (qr_max - logo_w) / 2,
        qr_y + qr_max + 10,
        logo_w,
        logo_h,
    )


def _draw_label_page(pdf, text_vis, qr_data, line_count):
    if line_count == 4:
        layout_profile = FOUR_LINE_LAYOUT
        text_y = landscape(A4)[1] - 100 - 56.6929 - 28.3465
    elif line_count == 5:
        layout_profile = FIVE_LINE_LAYOUT
        text_y = landscape(A4)[1] - 100 - 56.6929
    else:
        raise ValueError("Bitte ein gültiges Layout mit 4 oder 5 Zeilen auswählen.")

    width, height = landscape(A4)
    margin = SINGLE_LABEL_LAYOUT["margin"]
    font_sizes = layout_profile["font_sizes"]

    _draw_border(pdf, width, height, margin, SINGLE_LABEL_LAYOUT["border_width"])

    text_x = SINGLE_LABEL_LAYOUT["text_x"]
    left_width = width / 2 - 50
    lines = str(text_vis).split()

    if not lines:
        raise ValueError("Für das Schild fehlt der sichtbare Lagerplatztext.")

    for i, line in enumerate(lines):
        size = font_sizes[min(i, len(font_sizes) - 1)]
        pdf.setFont("Times-Bold", size)
        text_width = pdf.stringWidth(line, "Times-Bold", size)

        if line_count == 5 and (i == len(lines) - 1 or i == 4):
            text_y += 28.3465

        pdf.drawString(text_x + (left_width - text_width) / 2, text_y, line)
        text_y -= size + 22.6772

    _draw_footer(pdf)
    _draw_qr_and_logo(pdf, width, height, qr_data)


def fit_text_to_safe_area(pdf, lines, safe_area_width, safe_area_height, layout):
    if not lines:
        return layout["min_font_size"]

    font_size = layout["max_font_size"]

    while font_size >= layout["min_font_size"]:
        line_spacing = font_size * 1.2
        total_text_height = len(lines) * line_spacing
        longest_line_width = max(
            pdf.stringWidth(line, layout["font_name"], font_size)
            for line in lines
        )

        if longest_line_width <= safe_area_width and total_text_height <= safe_area_height:
            return font_size

        font_size -= 2

    return layout["min_font_size"]


def split_text_into_lines(pdf, text, safe_area_width, font_name, font_size):
    words = str(text).split()

    if not words:
        return []

    lines = []
    current_line = ""

    for word in words:
        test_line = word if current_line == "" else current_line + " " + word
        test_width = pdf.stringWidth(test_line, font_name, font_size)

        if test_width <= safe_area_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)
            current_line = word

    if current_line:
        lines.append(current_line)

    return lines


def generate_pdf_einzeln(text_vis, qr_data, output, zeilen):
    """Erstellt ein einzelnes Lagerplatzschild als PDF."""
    pdf = _create_pdf_canvas(output)
    _draw_label_page(pdf, text_vis, qr_data, zeilen)
    pdf.showPage()

    try:
        pdf.save()
    except PermissionError as exc:
        raise PermissionError(
            "Die PDF-Datei konnte nicht gespeichert werden. Bitte prüfe, ob die Datei "
            "bereits geöffnet ist oder ob du Schreibrechte im Zielordner hast."
        ) from exc


def generate_text_sign_pdf(title_text, text, output, show_safe_area=False):
    """Erstellt ein Textschild als PDF. Diese Funktion dient aktuell der Layoutentwicklung."""
    if not str(title_text).strip():
        raise ValueError("Für das Textschild fehlt der Titel.")

    if not str(text).strip():
        raise ValueError("Für das Textschild fehlt der Haupttext.")

    layout = TEXT_SIGN_A4_LANDSCAPE_LAYOUT
    pdf = _create_pdf_canvas(output)
    width, height = landscape(A4)

    margin = layout["margin"]
    _draw_border(pdf, width, height, margin, layout["border_width"])

    logo_img = _load_logo(layout["logo_path"])
    logo_w = layout["logo_max_width"]
    logo_h = logo_w * logo_img.size[1] / logo_img.size[0]
    logo_x = width - logo_w - layout["logo_padding_x"] - margin
    logo_y = height - logo_h - layout["logo_padding_y"] - margin

    pdf.drawImage(ImageReader(logo_img), logo_x, logo_y, logo_w, logo_h)

    pdf.setFont(layout["title_font_name"], layout["title_font_size"])

    title_x = margin + layout["title_area_padding"]
    title_y = (
        height
        - margin
        - layout["title_area_padding"]
        - layout["title_font_size"]
    )
    pdf.drawString(title_x, title_y, str(title_text).strip())

    safe_area_x = margin + layout["text_area_left_padding"]
    safe_area_y = margin + layout["text_area_bottom_padding"]
    safe_area_width = (
        width
        - (2 * margin)
        - layout["text_area_left_padding"]
        - layout["text_area_right_padding"]
    )
    safe_area_height = (
        height
        - (2 * margin)
        - layout["title_area_height"]
        - layout["text_area_top_padding"]
        - layout["text_area_bottom_padding"]
    )

    if show_safe_area:
        pdf.setStrokeColorRGB(1, 0, 0)
        pdf.rect(safe_area_x, safe_area_y, safe_area_width, safe_area_height)

    lines = split_text_into_lines(
        pdf,
        text,
        safe_area_width,
        layout["font_name"],
        layout["max_font_size"],
    )

    if not lines:
        raise ValueError("Für das Textschild konnte kein darstellbarer Text ermittelt werden.")

    font_size = fit_text_to_safe_area(
        pdf,
        lines,
        safe_area_width,
        safe_area_height,
        layout,
    )

    pdf.setFont(layout["font_name"], font_size)
    line_spacing = font_size * 1.2
    total_text_height = len(lines) * line_spacing
    start_y = safe_area_y + (safe_area_height + total_text_height) / 2 - font_size

    for line in lines:
        line_width = pdf.stringWidth(line, layout["font_name"], font_size)
        line_x = safe_area_x + (safe_area_width - line_width) / 2
        pdf.drawString(line_x, start_y, line)
        start_y -= line_spacing

    pdf.showPage()

    try:
        pdf.save()
    except PermissionError as exc:
        raise PermissionError(
            "Die Textschild-PDF konnte nicht gespeichert werden. Bitte prüfe, ob die Datei "
            "bereits geöffnet ist oder ob du Schreibrechte im Zielordner hast."
        ) from exc


def _load_workbook_for_batch(excel):
    excel_path = ensure_excel_input_path(excel)

    try:
        return openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as exc:
        raise ValueError(
            "Die Excel-Datei konnte nicht gelesen werden. Bitte eine gültige .xlsx-Datei auswählen."
        ) from exc
    except PermissionError as exc:
        raise PermissionError(
            "Die Excel-Datei konnte nicht geöffnet werden. Bitte prüfe, ob sie noch in Excel geöffnet ist."
        ) from exc


def _iter_valid_batch_rows(sheet):
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        visible_text = row[0] if len(row) >= 1 else None
        qr_data = row[1] if len(row) >= 2 else None

        if not visible_text or not qr_data:
            yield row_number, None, None
            continue

        yield row_number, str(visible_text).strip(), str(qr_data).strip()


def _generate_batch_pdf(excel, output, line_count):
    if line_count not in (4, 5):
        raise ValueError("Bitte ein gültiges Layout mit 4 oder 5 Zeilen auswählen.")

    wb = _load_workbook_for_batch(excel)
    pdf = _create_pdf_canvas(output)

    created_count = 0
    skipped_count = 0

    try:
        sheet = wb.active

        if sheet.max_column < 2:
            raise ValueError(
                "Die Excel-Datei muss mindestens zwei Spalten enthalten:\n"
                "Spalte A = sichtbarer Lagerplatztext, Spalte B = QR-Code-Inhalt."
            )

        for _row_number, visible_text, qr_data in _iter_valid_batch_rows(sheet):
            if visible_text is None or qr_data is None:
                skipped_count += 1
                continue

            _draw_label_page(pdf, visible_text, qr_data, line_count)
            pdf.showPage()
            created_count += 1

        if created_count == 0:
            raise ValueError(
                "In der Excel-Datei wurden keine gültigen Datensätze gefunden.\n"
                "Bitte prüfe Spalte A und B. Beide Spalten müssen je Zeile gefüllt sein."
            )

        pdf.save()
    except PermissionError as exc:
        raise PermissionError(
            "Die Sammel-PDF konnte nicht gespeichert werden. Bitte prüfe, ob die Datei "
            "bereits geöffnet ist oder ob du Schreibrechte im Zielordner hast."
        ) from exc
    finally:
        wb.close()

    return {
        "created": created_count,
        "skipped": skipped_count,
    }


def generate_batch_pdf_4(excel, output):
    return _generate_batch_pdf(excel, output, 4)


def generate_batch_pdf_5(excel, output):
    return _generate_batch_pdf(excel, output, 5)


def generate_batch_pdf(excel, output, line_count):
    return _generate_batch_pdf(excel, output, line_count)
