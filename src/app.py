import openpyxl
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from PIL import Image, ImageTk
from src.paths import resource_path
from src.pdf_generator import generate_pdf_einzeln
from src.pdf_generator import generate_text_sign_pdf
from src.validators import is_single_pdf_valid
from src.gui.release_notes_window import ReleaseNotesWindow
from src.gui.handbuch_window import HandbuchWindow
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import webbrowser
import urllib.parse


# ---- Farbdefinitionen ----
BG_COLOR = "#9d9d9d"
BUTTON_COLOR = "#00703c"
class QRCodeGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Lagerplatz-QRCode-Generator Version 1.0")
        self.root.geometry("810x600")
        self.root.resizable(False, False)
        self.root.configure(bg=BG_COLOR)

        # Tabs
        style = ttk.Style()
        style.theme_use('default')
        style.configure('TNotebook.Tab', background=BG_COLOR, font=('Arial', 11, 'bold'), padding=[16, 6])
        style.map('TNotebook.Tab', background=[('selected', '#bdbdbd')])

        self.notebook = ttk.Notebook(root)
        self.tab_single = tk.Frame(self.notebook, bg=BG_COLOR)
        self.tab_batch = tk.Frame(self.notebook, bg=BG_COLOR)
        self.notebook.add(self.tab_single, text="Einzelerstellung")
        self.notebook.add(self.tab_batch, text="Sammelverarbeitung")
        self.notebook.pack(fill="both", expand=True, padx=40, pady=28)

        # Trennstrich unter den Tabs
        separator = tk.Frame(root, bg="#6e6e6e", height=2)
        separator.pack(fill="x", padx=30, pady=(0, 8))

        self.notebook.select(0)  # Einzelerstellung ist Standard

        # Einzel-TAB Variablen
        self.single_lagerplatz = tk.StringVar()
        self.single_layout = tk.StringVar(value="Bitte wählen...")
        self.single_output = tk.StringVar()
        self.single_status = None
        self.single_btn_pdf = None

        # Batch-TAB Variablen
        self.batch_excel_path = tk.StringVar()
        self.batch_layout = tk.StringVar(value="Bitte wählen...")
        self.batch_output_path = tk.StringVar()
        self.batch_status = None
        self.batch_btn_pdf = None

        self.build_tab_single()
        self.build_tab_batch()
        self.build_version_label()

    # ==== VERSION LABEL MIT RELEASE NOTES ====
    def build_version_label(self):
        frame = tk.Frame(self.root, bg=BG_COLOR)
        frame.place(relx=0, rely=1.0, anchor="sw", x=12, y=-8)

        self.version_lbl = tk.Label(
            frame,
            text="© Fabian Wiese – Version 1.0 ⓘ",
            fg="black",
            bg=BG_COLOR,
            cursor="arrow"
        )
        self.version_lbl.pack(anchor="sw")
        self.version_lbl.bind("<Enter>", self._on_version_hover)
        self.version_lbl.bind("<Leave>", self._on_version_leave)
        self.version_lbl.bind("<Button-1>", self.show_release_notes)

    def _on_version_hover(self, event):
        self.version_lbl.config(fg="white", font=("Arial", 10, "underline"), cursor="hand2")

    def _on_version_leave(self, event):
        self.version_lbl.config(fg="black", font=("Arial", 10, "normal"), cursor="arrow")

    def show_release_notes(self, event=None):
        ReleaseNotesWindow(self.root)

    # ==== EINZELERSTELLUNG ====
    def build_tab_single(self):
        frame = self.tab_single

        # Firmenlogo
        logo_path = resource_path("Schnellecke Logistics_transparent.png")
        if os.path.exists(logo_path):
            img = Image.open(logo_path)
            img = img.resize((250, 80), Image.LANCZOS)
            logo_img = ImageTk.PhotoImage(img)
            logo_label = tk.Label(frame, image=logo_img, bg=BG_COLOR)
            logo_label.image = logo_img
            logo_label.place(relx=1.0, rely=0.0, anchor="ne")

        # Titel
        tk.Label(
            frame,
            text="Ein einzelnes Lagerplatzschild generieren",
            font=("Arial", 12, "bold"),
            bg=BG_COLOR
        ).grid(row=0, column=0, sticky="w", pady=(10, 8), columnspan=2)

        # Layout Dropdown
        tk.Label(
            frame,
            text="Wie viele Zeilen Text soll dein Lagerplatzetikett enthalten?",
            font=("Arial", 10),
            bg=BG_COLOR
        ).grid(row=1, column=0, sticky="w", padx=(0,0))
        layout_dropdown = ttk.Combobox(
            frame,
            textvariable=self.single_layout,
            values=["Bitte wählen...", "4 Zeilen pro PDF-Seite", "5 Zeilen pro PDF-Seite"],
            state="readonly",
            width=30
        )
        layout_dropdown.grid(row=2, column=0, sticky="w", pady=(0, 10))
        layout_dropdown.bind("<<ComboboxSelected>>", lambda e: self.update_single_button())

        # Eingabe Lagerplatz (systemisch)
        tk.Label(
            frame,
            text="Lagerplatz (systemisch, mit Bindestrichen):",
            font=("Arial", 10),
            bg=BG_COLOR
        ).grid(row=3, column=0, sticky="w", pady=(0, 2))
        tk.Entry(frame, textvariable=self.single_lagerplatz, width=50, bg=BG_COLOR, fg="black", insertbackground="black").grid(row=4, column=0, sticky="w")

        # Speicherort Auswahl
        tk.Label(
            frame,
            text="Speicherort für das PDF:",
            font=("Arial", 10),
            bg=BG_COLOR
        ).grid(row=5, column=0, pady=(14, 0), sticky="w")
        tk.Entry(frame, textvariable=self.single_output, width=60, bg=BG_COLOR, fg="black", insertbackground="black").grid(row=6, column=0, sticky="w")
        tk.Button(frame, text="Durchsuchen", command=self.single_save_pdf, bg=BUTTON_COLOR, fg="black").grid(row=6, column=1, padx=10)

        # Status-Label
        self.single_status = tk.Label(frame, text="Bereit", fg="blue", bg=BG_COLOR)
        self.single_status.grid(row=7, column=0, pady=(15, 0), sticky="w")

        # Buttons
        btn_frame = tk.Frame(frame, bg=BG_COLOR)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=20, sticky="w")

        self.single_btn_pdf = tk.Button(
            btn_frame,
            text="PDF erstellen",
            command=self.on_single_generate,
            bg=BUTTON_COLOR,
            fg="black",
            padx=20,
            pady=5,
            state="disabled"
        )
        self.single_btn_pdf.pack(side="left", padx=(0, 20))

        tk.Button(
            btn_frame,
            text="Kontakt / Support",
            command=self.open_email,
            bg=BUTTON_COLOR,
            fg="black",
            padx=18,
            pady=5
        ).pack(side="left")

        # Handbuch-Button unten rechts
        tk.Button(
            frame,
            text="Handbuch",
            command=self.show_handbuch,
            bg=BUTTON_COLOR,
            fg="black",
            relief="raised"
        ).place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)
        
        tk.Button(
            btn_frame,
            text="Textschild Test",
            command=lambda: generate_text_sign_pdf(
                "TESTSCHILD",
                "PARKPLATZ FÜR BESUCHER UND LIEFERANTEN SOWIE EXTERNE DIENSTLEISTER",
                "output/test_textschild.pdf",
                True
            ),
            bg=BUTTON_COLOR,
            fg="black",
            padx=18,
            pady=5
        ).pack(side="left", padx=(0, 20))

        # Live-Validation
        frame.bind_all('<KeyRelease>', lambda e: self.update_single_button())
        frame.bind_all('<<ComboboxSelected>>', lambda e: self.update_single_button())

    def show_handbuch(self):
        HandbuchWindow(self.root)

    def update_single_button(self):
        lagerplatz = self.single_lagerplatz.get().strip()
        pdf = self.single_output.get().strip()
        layout = self.single_layout.get()
        if is_single_pdf_valid(lagerplatz, pdf, layout):
            self.single_btn_pdf.config(state="normal")
        else:
            self.single_btn_pdf.config(state="disabled")

    def single_save_pdf(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF-Dateien", "*.pdf")]
        )
        if path:
            self.single_output.set(path)
            self.update_single_button()

    def on_single_generate(self):
        self.single_status.config(text="PDF wird erstellt...")
        self.root.update_idletasks()
        choice = 5 if "5 Zeilen" in self.single_layout.get() else 4
        lagerplatz_sys = self.single_lagerplatz.get().strip()
        lagerplatz_vis = lagerplatz_sys.replace("-", " ")
        qr_data = lagerplatz_sys

        output = self.single_output.get().strip()
        if choice == 4:
            generate_pdf_einzeln(lagerplatz_vis, qr_data, output, 4)
        else:
            generate_pdf_einzeln(lagerplatz_vis, qr_data, output, 5)
        self.single_status.config(text="Fertig")
        messagebox.showinfo("Fertig", f"PDF erfolgreich erstellt:\n{output}")

    # ==== SAMMELVERARBEITUNG ====
    def build_tab_batch(self):
        frame = self.tab_batch

        # Firmenlogo
        logo_path = resource_path("Schnellecke Logistics_transparent.png")
        if os.path.exists(logo_path):
            img = Image.open(logo_path)
            img = img.resize((250, 80), Image.LANCZOS)
            logo_img = ImageTk.PhotoImage(img)
            logo_label = tk.Label(frame, image=logo_img, bg=BG_COLOR)
            logo_label.image = logo_img
            logo_label.place(relx=1.0, rely=0.0, anchor="ne")

        # Titel
        tk.Label(
            frame,
            text="Lagerplatz-Schilder aus Excel-Datei generieren.",
            font=("Arial", 12, "bold"),
            bg=BG_COLOR
        ).grid(row=0, column=0, sticky="w", pady=(10, 8), columnspan=2)

        # Layout Dropdown
        tk.Label(
            frame,
            text="Wie viele Zeilen Text soll dein Lagerplatzetikett enthalten?",
            font=("Arial", 10),
            bg=BG_COLOR
        ).grid(row=1, column=0, sticky="w", padx=(0,0))
        layout_dropdown = ttk.Combobox(
            frame,
            textvariable=self.batch_layout,
            values=["Bitte wählen...", "4 Zeilen pro PDF-Seite", "5 Zeilen pro PDF-Seite"],
            state="readonly",
            width=30
        )
        layout_dropdown.grid(row=2, column=0, sticky="w", pady=(0, 14))
        layout_dropdown.bind("<<ComboboxSelected>>", lambda e: self.update_batch_button())

        # Excel-Dateiauswahl
        tk.Label(frame, text="Excel-Datei auswählen:", bg=BG_COLOR).grid(row=3, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.batch_excel_path, width=60, bg=BG_COLOR, fg="black", insertbackground="black").grid(row=4, column=0, sticky="w")
        tk.Button(frame, text="Durchsuchen", command=self.browse_batch_excel, bg=BUTTON_COLOR, fg="black").grid(row=4, column=1, padx=10)

        # Hinweistext
        tk.Label(
            frame, text="• Es muss eine .xlsx-Datei ohne Überschrift verwendet werden.", font=("Arial", 9), fg="darkblue", bg=BG_COLOR
        ).grid(row=5, column=0, sticky="w", pady=(12, 0))
        tk.Label(
            frame, text="• Spalte A: Lagerplatz mit Leerzeichen (z. B. 05 H2 11 093 L)", font=("Arial", 9), fg="darkblue", bg=BG_COLOR
        ).grid(row=6, column=0, sticky="w")
        tk.Label(
            frame, text="• Spalte B: Lagerplatz mit Bindestrichen (z. B. 05-H2-11-093-L)", font=("Arial", 9), fg="darkblue", bg=BG_COLOR
        ).grid(row=7, column=0, sticky="w")

        # Speicherort Auswahl
        tk.Label(frame, text="Speicherort für die kombinierte PDF:", bg=BG_COLOR).grid(row=8, column=0, pady=(16, 0), sticky="w")
        tk.Entry(frame, textvariable=self.batch_output_path, width=60, bg=BG_COLOR, fg="black", insertbackground="black").grid(row=9, column=0, sticky="w")
        tk.Button(frame, text="Durchsuchen", command=self.save_batch_pdf, bg=BUTTON_COLOR, fg="black").grid(row=9, column=1, padx=10)

        # Status-Label
        self.batch_status = tk.Label(frame, text="Bereit", fg="blue", bg=BG_COLOR)
        self.batch_status.grid(row=10, column=0, pady=(15, 0), sticky="w")

        # Buttons
        btn_frame = tk.Frame(frame, bg=BG_COLOR)
        btn_frame.grid(row=11, column=0, columnspan=2, pady=20, sticky="w")

        self.batch_btn_pdf = tk.Button(
            btn_frame,
            text="PDF erstellen",
            command=self.on_batch_generate,
            bg=BUTTON_COLOR,
            fg="black",
            padx=20,
            pady=5,
            state="disabled"
        )
        self.batch_btn_pdf.pack(side="left", padx=(0, 20))

        tk.Button(
            btn_frame,
            text="Kontakt / Support",
            command=self.open_email,
            bg=BUTTON_COLOR,
            fg="black",
            padx=18,
            pady=5
        ).pack(side="left")

        # Handbuch-Button unten rechts
        tk.Button(
            frame,
            text="Handbuch",
            command=self.show_handbuch,
            bg=BUTTON_COLOR,
            fg="black",
            relief="raised"
        ).place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)

        # Live-Validation
        frame.bind_all('<KeyRelease>', lambda e: self.update_batch_button())
        frame.bind_all('<<ComboboxSelected>>', lambda e: self.update_batch_button())

    def update_batch_button(self):
        excel = self.batch_excel_path.get().strip()
        pdf = self.batch_output_path.get().strip()
        zeilen = self.batch_layout.get()
        if (zeilen in ["4 Zeilen pro PDF-Seite", "5 Zeilen pro PDF-Seite"]
            and excel.endswith(".xlsx")
            and pdf.endswith(".pdf")
            and os.path.exists(excel)):
            self.batch_btn_pdf.config(state="normal")
        else:
            self.batch_btn_pdf.config(state="disabled")

    def browse_batch_excel(self):
        path = filedialog.askopenfilename(
            title="Excel-Datei wählen",
            filetypes=[("Excel-Dateien", "*.xlsx")]
        )
        if path:
            self.batch_excel_path.set(path)
            self.update_batch_button()

    def save_batch_pdf(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF-Dateien", "*.pdf")]
        )
        if path:
            self.batch_output_path.set(path)
            self.update_batch_button()

    def on_batch_generate(self):
        self.batch_status.config(text="PDF wird erstellt...")
        self.root.update_idletasks()
        choice = 5 if "5 Zeilen" in self.batch_layout.get() else 4
        if choice == 4:
            self.generate_pdf_4()
        else:
            self.generate_pdf_5()
        self.batch_status.config(text="Fertig")
        messagebox.showinfo("Fertig", f"PDF erfolgreich erstellt:\n{self.batch_output_path.get()}")

    def open_email(self):
        empfaenger = "fabian.wiese@schnellecke.com"
        betreff = "Supportanfrage für QR-Code Generator"
        body = (
            "Name:\n"
            "Vorname:\n"
            "Telefonnummer (optional):\n"
            "E-Mail-Adresse:\n"
            "Beschreibung / Verbesserungsvorschlag:\n\n"
            "Falls möglich, bitte ein Screenshot-Bild vom Fehler mit anhängen."
        )
        mailto_link = (
            f"mailto:{empfaenger}?subject={urllib.parse.quote(betreff)}&body={urllib.parse.quote(body)}"
        )
        webbrowser.open(mailto_link)

    # ==== BATCH GENERIERUNG (wie bisher) ====
    def generate_pdf_4(self):
        excel = self.batch_excel_path.get()
        output = self.batch_output_path.get()
        wb = openpyxl.load_workbook(excel)
        sheet = wb.active
        font_sizes = [60, 70, 75, 105]
        border_color = colors.Color(0/255,112/255,60/255)
        pdf = canvas.Canvas(output, pagesize=landscape(A4))
        width, height = landscape(A4)
        for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
            pdf.setStrokeColor(border_color)
            pdf.setLineWidth(14.1732)
            margin = 28.3465*2
            pdf.rect(margin, margin, width-2*margin, height-2*margin)
            text_y = height-100-56.6929-28.3465
            text_x = 50
            left_width = width/2-50
            if row[0]:
                lines = str(row[0]).split()
                for i, line in enumerate(lines):
                    size = font_sizes[min(i,len(font_sizes)-1)]
                    pdf.setFont("Times-Bold", size)
                    tw = pdf.stringWidth(line, "Times-Bold", size)
                    pdf.drawString(text_x+(left_width-tw)/2, text_y, line)
                    text_y -= size+22.6772
            footer_text = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.0"
            pdf.setFont("Times-Roman",12)
            pdf.drawString(50,30, footer_text)
            qr_data = row[1] if row[1] else "https://example.com"
            qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=1)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            tmp_qr = f"temp_qr_{idx}.png"; qr_img.save(tmp_qr)
            qr_max = min(width/2-50, height-200)
            qr_img = Image.open(tmp_qr).resize((int(qr_max),int(qr_max)), Image.LANCZOS)
            qr_x = width/2+50-42.5197-56.6929; qr_y = (height-qr_max)/2-28.3465
            pdf.drawImage(tmp_qr, qr_x, qr_y, qr_max, qr_max)
            os.remove(tmp_qr)
            logo_file = resource_path("assets/Logo_White.jpeg")
            logo_img = Image.open(logo_file)
            logo_w=qr_max*0.5; logo_h=logo_w*logo_img.size[1]/logo_img.size[0]
            logo_img=logo_img.resize((int(logo_w),int(logo_h)), Image.LANCZOS)
            tmp_logo=f"temp_logo_{idx}.png"; logo_img.save(tmp_logo)
            pdf.drawImage(tmp_logo, qr_x+(qr_max-logo_w)/2, qr_y+qr_max+10, logo_w, logo_h)
            os.remove(tmp_logo)
            pdf.showPage()
        pdf.save()

    def generate_pdf_5(self):
        excel = self.batch_excel_path.get()
        output = self.batch_output_path.get()
        wb = openpyxl.load_workbook(excel)
        sheet = wb.active
        font_sizes = [40, 50, 75, 105, 80]
        border_color = colors.Color(0/255,112/255,60/255)
        pdf = canvas.Canvas(output, pagesize=landscape(A4))
        width, height = landscape(A4)
        for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
            pdf.setStrokeColor(border_color)
            pdf.setLineWidth(14.1732)
            margin = 28.3465*2
            pdf.rect(margin, margin, width-2*margin, height-2*margin)
            text_y = height-100-56.6929
            text_x = 50
            left_width = width/2-50
            if row[0]:
                lines = str(row[0]).split()
                for i, line in enumerate(lines):
                    size = font_sizes[min(i,len(font_sizes)-1)]
                    pdf.setFont("Times-Bold", size)
                    tw = pdf.stringWidth(line, "Times-Bold", size)
                    if i==len(lines)-1 or i==4:
                        text_y += 28.3465
                    pdf.drawString(text_x+(left_width-tw)/2, text_y, line)
                    text_y -= size+22.6772
            footer_text = "Erstellt: Fabian Wiese, generiert durch Lagerplatz-QRCode-Generator Version 1.0"
            pdf.setFont("Times-Roman",12)
            pdf.drawString(50,30, footer_text)
            qr_data = row[1] if row[1] else "https://example.com"
            qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=1)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            tmp_qr = f"temp_qr_{idx}.png"; qr_img.save(tmp_qr)
            qr_max = min(width/2-50, height-200)
            qr_img = Image.open(tmp_qr).resize((int(qr_max),int(qr_max)), Image.LANCZOS)
            qr_x = width/2+50-42.5197-56.6929; qr_y = (height-qr_max)/2-28.3465
            pdf.drawImage(tmp_qr, qr_x, qr_y, qr_max, qr_max)
            os.remove(tmp_qr)
            logo_file = resource_path("assets/Logo_White.jpeg")
            logo_img = Image.open(logo_file)
            logo_w=qr_max*0.5; logo_h=logo_w*logo_img.size[1]/logo_img.size[0]
            logo_img=logo_img.resize((int(logo_w),int(logo_h)), Image.LANCZOS)
            tmp_logo=f"temp_logo_{idx}.png"; logo_img.save(tmp_logo)
            pdf.drawImage(tmp_logo, qr_x+(qr_max-logo_w)/2, qr_y+qr_max+10, logo_w, logo_h)
            os.remove(tmp_logo)
            pdf.showPage()
        pdf.save()

def main():
    root = tk.Tk()
    app = QRCodeGeneratorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()